from datetime import date, datetime
from typing import Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import openpyxl
import re

from app.db.models.schedule import DailyDispatchSchedule, HourlyDemand, HourlySupply
from app.db.models.baseload import BaseloadPlant
from app.services.baseload_service import BaseloadService
from app.core.logging import get_logger

logger = get_logger(__name__)

DEMAND_ENTITY_KEYWORDS = {
    "total ecg demand": "ECG",
    "nedco": "NEDCo",
    "valco": "VALCO",
    "mines": "Mines",
    "export": "Export",
}

EMBEDDED_SOURCE_KEYWORDS = {
    "trojan": "Trojan I (Tema)",
    "meienergy": "Meienergy",
    "bxc solar": "BXC Solar",
    "bxc": "BXC Solar",
}

TOTAL_DEMAND_KEYWORDS = ["scheduled demand from the nits", "total demand", "nits"]


class ScheduleService:

    async def parse_and_store(
        self,
        db: AsyncSession,
        filepath: str,
        filename: str,
        schedule_date: Optional[date] = None,
    ) -> DailyDispatchSchedule:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[wb.sheetnames[0]]

        header_info = self._extract_header_info(ws)
        dispatch_date = schedule_date or header_info.get("date") or date.today()

        existing = await db.execute(
            select(DailyDispatchSchedule).where(
                DailyDispatchSchedule.date == dispatch_date
            )
        )
        schedule_obj = existing.scalar_one_or_none()
        if schedule_obj:
            logger.warning(
                "Schedule for %s already exists (id=%d), deleting old rows",
                dispatch_date,
                schedule_obj.id,
            )
            await db.execute(
                HourlyDemand.__table__.delete().where(
                    HourlyDemand.schedule_id == schedule_obj.id
                )
            )
            await db.execute(
                HourlySupply.__table__.delete().where(
                    HourlySupply.schedule_id == schedule_obj.id
                )
            )

        if not schedule_obj:
            schedule_obj = DailyDispatchSchedule(
                date=dispatch_date,
                status="draft",
                source_filename=filename,
            )
            db.add(schedule_obj)
            await db.flush()
            await db.refresh(schedule_obj)

        hour_row_idx, hour_columns = self._find_hour_row(ws)
        if hour_row_idx is None:
            raise ValueError("Could not find 'Hour' header row in the Excel sheet")

        demand_rows = self._parse_demand_rows(ws, hour_row_idx, hour_columns)
        supply_rows = self._parse_supply_rows(ws, hour_row_idx, hour_columns)

        for entry in demand_rows:
            record = HourlyDemand(
                schedule_id=schedule_obj.id,
                hour=entry["hour"],
                entity_name=entry["entity_name"],
                demand_mw=entry["demand_mw"],
                is_forecasted=entry.get("is_forecasted", False),
            )
            db.add(record)

        for entry in supply_rows:
            record = HourlySupply(
                schedule_id=schedule_obj.id,
                hour=entry["hour"],
                plant_name=entry["plant_name"],
                supply_mw=entry["supply_mw"],
            )
            db.add(record)

        await db.commit()
        await db.refresh(schedule_obj)

        await self._merge_baseload_plants(db, schedule_obj.id)
        await db.commit()

        logger.info(
            "Stored schedule %d (%s): %d demand rows, %d supply rows",
            schedule_obj.id,
            dispatch_date,
            len(demand_rows),
            len(supply_rows),
        )

        wb.close()
        return schedule_obj

    def _extract_header_info(self, ws):
        info = {}
        for row in ws.iter_rows(min_row=1, max_row=40, values_only=True):
            raw = [c for c in row if c is not None]
            if not raw:
                continue
            text = " ".join(str(c) for c in raw if not isinstance(c, datetime)).lower()
            if "dispatch day" in text or "details of demand" in text:
                for c in raw:
                    if isinstance(c, datetime):
                        info["date"] = c.date()
                        break
            if "organisation" in text and len(raw) > 2:
                info["organisation"] = str(raw[2]).strip()
        return info

    def _find_hour_row(self, ws):
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
            first = str(row[1] if len(row) > 1 else "").strip().lower() if row[0] is None else str(row[0]).strip().lower()
            if first == "" and len(row) > 1:
                first = str(row[1]).strip().lower()
            if first in ("hour", "hr", "he"):
                cols = {}
                for col_idx, val in enumerate(row):
                    if val is not None and str(val).strip().isdigit():
                        cols[int(str(val).strip())] = col_idx
                if cols:
                    return row_idx, cols
        return None, None

    def _parse_demand_rows(self, ws, hour_row_idx, hour_columns):
        results = []
        for row in ws.iter_rows(min_row=hour_row_idx + 1, max_row=ws.max_row, values_only=True):
            label = str(row[1] if len(row) > 1 else "").strip() if len(row) > 1 else ""
            if not label or label == "None":
                continue

            label_lower = label.lower()

            matched_entity = None
            is_total = False
            for keyword, entity in DEMAND_ENTITY_KEYWORDS.items():
                if keyword in label_lower:
                    matched_entity = entity
                    break

            for keyword in TOTAL_DEMAND_KEYWORDS:
                if keyword in label_lower:
                    matched_entity = "NITS_Total"
                    is_total = True
                    break

            if matched_entity is None:
                continue

            for hour, col_idx in hour_columns.items():
                if col_idx < len(row):
                    val = row[col_idx]
                    if val is not None:
                        try:
                            mw = float(val)
                            results.append({
                                "hour": hour,
                                "entity_name": matched_entity,
                                "demand_mw": round(mw, 2),
                                "is_forecasted": False,
                            })
                        except (ValueError, TypeError):
                            pass

        return results

    def _parse_supply_rows(self, ws, hour_row_idx, hour_columns):
        results = []
        for row in ws.iter_rows(min_row=hour_row_idx + 1, max_row=ws.max_row, values_only=True):
            label_raw = str(row[1] if len(row) > 1 else "").strip() if len(row) > 1 else ""
            if not label_raw or label_raw == "None":
                continue

            label_lower = label_raw.lower()

            matched_plant = None
            for keyword, plant in EMBEDDED_SOURCE_KEYWORDS.items():
                if keyword in label_lower:
                    matched_plant = plant
                    break

            if matched_plant is None:
                continue

            for hour, col_idx in hour_columns.items():
                if col_idx < len(row):
                    val = row[col_idx]
                    if val is not None:
                        try:
                            mw = float(val)
                            results.append({
                                "hour": hour,
                                "plant_name": matched_plant,
                                "supply_mw": round(mw, 2),
                            })
                        except (ValueError, TypeError):
                            pass

        return results

    async def _merge_baseload_plants(self, db: AsyncSession, schedule_id: int) -> None:
        """Insert HourlySupply rows for baseload plants not already in the schedule."""
        existing_result = await db.execute(
            select(HourlySupply.plant_name).where(
                HourlySupply.schedule_id == schedule_id
            ).distinct()
        )
        existing_plants = {row[0] for row in existing_result.all()}

        baseload_service = BaseloadService()
        baseload_plants = await baseload_service.get_all_active(db)

        inserted = 0
        for bp in baseload_plants:
            label = baseload_service.get_plant_label(bp)
            if label in existing_plants:
                continue

            for hour in range(1, 25):
                record = HourlySupply(
                    schedule_id=schedule_id,
                    hour=hour,
                    plant_name=label,
                    supply_mw=bp.constant_mw,
                )
                db.add(record)
                inserted += 1

        if inserted:
            await db.flush()
            logger.info(f"Merged {inserted} baseload supply rows for schedule {schedule_id}")

    async def get_schedule(self, db: AsyncSession, schedule_id: int) -> Optional[dict]:
        result = await db.execute(
            select(DailyDispatchSchedule).where(DailyDispatchSchedule.id == schedule_id)
        )
        schedule_obj = result.scalar_one_or_none()
        if not schedule_obj:
            return None

        demand_result = await db.execute(
            select(HourlyDemand).where(HourlyDemand.schedule_id == schedule_id)
        )
        demand_rows = demand_result.scalars().all()

        supply_result = await db.execute(
            select(HourlySupply).where(HourlySupply.schedule_id == schedule_id)
        )
        supply_rows = supply_result.scalars().all()

        baseload_service = BaseloadService()
        baseload_plants = await baseload_service.get_all_active(db)
        baseload_labels = {baseload_service.get_plant_label(bp) for bp in baseload_plants}
        baseload_categories = {baseload_service.get_plant_label(bp): bp.category for bp in baseload_plants}

        return {
            "id": schedule_obj.id,
            "date": schedule_obj.date,
            "status": schedule_obj.status,
            "source_filename": schedule_obj.source_filename,
            "operator_notes": schedule_obj.operator_notes,
            "created_at": schedule_obj.created_at,
            "updated_at": schedule_obj.updated_at,
            "demand": [
                {
                    "hour": d.hour,
                    "entity_name": d.entity_name,
                    "demand_mw": d.demand_mw,
                    "is_forecasted": d.is_forecasted,
                }
                for d in demand_rows
            ],
            "supply": [
                {
                    "hour": s.hour,
                    "plant_name": s.plant_name,
                    "supply_mw": s.supply_mw,
                    "is_baseload": s.plant_name in baseload_labels,
                    "category": baseload_categories.get(s.plant_name, ""),
                }
                for s in supply_rows
            ],
        }
